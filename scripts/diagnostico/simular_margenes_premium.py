#!/usr/bin/env python
"""
scripts/diagnostico/simular_margenes_premium.py

Simula 3 escenarios de margen mínimo (22% / 15% / 10%) sobre la franja de
ALTO COSTO, usando las medianas ya limpias (post re-espía v7 + recalibración).

Para cada producto alto costo (costo neto > UMBRAL_ALTO_COSTO) calcula el
precio que daría el motor con cada factor de piso, lo compara con el precio
actual de Shopify, y clasifica:
  - Competitivo: el mercado está sobre tu piso rentable → puedes seguirlo.
  - No competitivo: el mercado vende ≤ tu piso rentable → quedas en el piso
    (sobre el mercado). Bajar el margen no te hace el más barato porque tu
    costo no da.
  - Monopolio: sin datos de mercado.

Genera un Excel comparativo (NO toca Shopify ni cambia precios).

Uso:
    python scripts/diagnostico/simular_margenes_premium.py
"""

import os
import sys
import json
import math
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
sys.path.insert(0, BASE_DIR)

ARCHIVO_MEDIVEN = os.path.join(BASE_DIR, "mediven_full.json")
ARCHIVO_MERCADO = os.path.join(BASE_DIR, "data", "precios_mercado.json")
ARCHIVO_SHOPIFY = os.path.join(BASE_DIR, "shopify_full.json")
DIR_OUT = os.path.join(BASE_DIR, "data")

# === Parámetros del motor (deben coincidir con precios.py) ===
COMISION_TOTAL = float(os.getenv("COMISION_MP", "0.045")) + float(os.getenv("COMISION_SHOPIFY", "0.02"))
FACTOR_MONOPOLIO = float(os.getenv("FACTOR_MONOPOLIO", "1.50"))
MARKUP_MAX = float(os.getenv("MARKUP_MAX", "3.5"))
AGRESIVIDAD_ALTO_COSTO = float(os.getenv("AGRESIVIDAD_ALTO_COSTO", "1.00"))
UMBRAL_ALTO_COSTO = int(os.getenv("UMBRAL_ALTO_COSTO", "40000"))  # costo NETO

# Escenarios de margen mínimo a simular (factor de piso)
ESCENARIOS = [("Margen 22%", 1.22), ("Margen 15%", 1.15), ("Margen 10%", 1.10)]


def cargar_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def redondear_bonito(precio):
    if precio <= 0:
        return 0
    return int(math.ceil(precio / 100.0) * 100)


def precio_motor(costo_neto, mediana, factor_piso):
    """Replica la lógica del motor (alto costo) para un factor de piso dado.
    Devuelve (precio, clasificacion)."""
    costo_iva = costo_neto * 1.19
    precio_piso = (costo_iva * factor_piso) / (1 - COMISION_TOTAL)
    precio_techo = costo_iva * MARKUP_MAX

    if not mediana:
        precio = min((costo_iva * FACTOR_MONOPOLIO) / (1 - COMISION_TOTAL), precio_techo)
        return redondear_bonito(precio), "Monopolio"

    if mediana <= precio_piso:
        precio = min(precio_piso, precio_techo)
        return redondear_bonito(precio), "No competitivo"

    precio = max(precio_piso, mediana * AGRESIVIDAD_ALTO_COSTO)
    precio = min(precio, precio_techo)
    return redondear_bonito(precio), "Competitivo"


def margen_real_pct(precio, costo_neto):
    if precio <= 0:
        return 0.0
    costo_iva = costo_neto * 1.19
    ingreso_neto = precio * (1 - COMISION_TOTAL)
    return (ingreso_neto - costo_iva) / costo_iva * 100


def construir_indice_shopify(shopify_data):
    """Mapa sku -> precio actual, robusto a distintas formas del cache."""
    idx = {}
    items = shopify_data if isinstance(shopify_data, list) else shopify_data.get("products", [])
    for prod in items:
        variants = prod.get("variants") or [prod]
        for v in variants:
            sku = str(v.get("sku") or prod.get("sku") or "").strip()
            if not sku:
                continue
            precio = v.get("price") or prod.get("price") or 0
            try:
                idx[sku] = float(precio)
            except (TypeError, ValueError):
                idx[sku] = 0.0
    return idx


def main():
    productos = cargar_json(ARCHIVO_MEDIVEN)
    mercado = cargar_json(ARCHIVO_MERCADO) if os.path.exists(ARCHIVO_MERCADO) else {}
    shopify_idx = {}
    if os.path.exists(ARCHIVO_SHOPIFY):
        shopify_idx = construir_indice_shopify(cargar_json(ARCHIVO_SHOPIFY))
    else:
        print("⚠️  No hay shopify_full.json; la columna 'Precio actual' saldrá vacía.")

    # Filtrar alto costo (costo neto > umbral)
    filas = []
    for p in productos:
        sku = str(p.get("Codigo", "")).strip()
        if not sku:
            continue
        try:
            costo_neto = float(p.get("Precio", 0) or 0)
        except (TypeError, ValueError):
            costo_neto = 0
        if costo_neto <= UMBRAL_ALTO_COSTO:
            continue

        reg = mercado.get(sku) or {}
        dm = reg.get("datos_mercado")
        mediana = dm.get("mediana_competitiva") if dm else None
        fuentes = dm.get("fuentes_validas") if dm else 0
        precio_actual = shopify_idx.get(sku, 0)

        fila = {
            "sku": sku,
            "desc": str(p.get("Descripcion", ""))[:50],
            "costo_neto": costo_neto,
            "costo_iva": costo_neto * 1.19,
            "mediana": mediana or 0,
            "fuentes": fuentes,
            "precio_actual": precio_actual,
        }
        for nombre_esc, factor in ESCENARIOS:
            precio, clasif = precio_motor(costo_neto, mediana, factor)
            fila[nombre_esc] = precio
            fila[nombre_esc + "_clasif"] = clasif
            fila[nombre_esc + "_margen"] = margen_real_pct(precio, costo_neto)
        filas.append(fila)

    if not filas:
        print("No se encontraron productos de alto costo. Revisa UMBRAL_ALTO_COSTO.")
        return

    # === Construir Excel ===
    wb = openpyxl.Workbook()
    azul = PatternFill("solid", fgColor="1F4E78")
    gris = PatternFill("solid", fgColor="D9D9D9")
    verde = PatternFill("solid", fgColor="C6EFCE")
    rojo = PatternFill("solid", fgColor="FFC7CE")
    amar = PatternFill("solid", fgColor="FFEB9C")
    blanco_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    # --- Hoja Resumen ---
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "SIMULACIÓN DE MÁRGENES — FRANJA ALTO COSTO (datos limpios v7)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Productos alto costo (costo neto > ${UMBRAL_ALTO_COSTO:,}): {len(filas)}"
    ws["A3"] = f"Generado: {datetime.now():%Y-%m-%d %H:%M}"

    encabezados = ["Escenario", "Competitivos", "No competitivos", "Monopolio",
                   "Margen total $ (vs costo)", "Suma precios", "Bajan vs hoy", "Suben vs hoy"]
    fila0 = 5
    for c, h in enumerate(encabezados, 1):
        cell = ws.cell(fila0, c, h)
        cell.fill = azul
        cell.font = blanco_bold
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, (nombre_esc, factor) in enumerate(ESCENARIOS):
        r = fila0 + 1 + i
        comp = sum(1 for f in filas if f[nombre_esc + "_clasif"] == "Competitivo")
        noc = sum(1 for f in filas if f[nombre_esc + "_clasif"] == "No competitivo")
        mono = sum(1 for f in filas if f[nombre_esc + "_clasif"] == "Monopolio")
        margen_total = sum((f[nombre_esc] * (1 - COMISION_TOTAL) - f["costo_iva"]) for f in filas)
        suma = sum(f[nombre_esc] for f in filas)
        bajan = sum(1 for f in filas if f["precio_actual"] and f[nombre_esc] < f["precio_actual"])
        suben = sum(1 for f in filas if f["precio_actual"] and f[nombre_esc] > f["precio_actual"])
        valores = [nombre_esc, comp, noc, mono, int(margen_total), int(suma), bajan, suben]
        for c, v in enumerate(valores, 1):
            ws.cell(r, c, v)

    # Nota interpretativa
    nr = fila0 + len(ESCENARIOS) + 3
    ws.cell(nr, 1, "Cómo leer esto:").font = bold
    notas = [
        "• Competitivos: el mercado está sobre tu piso → el precio sigue al mercado. Aquí el margen elegido SÍ cambia el precio.",
        "• No competitivos: el mercado vende ≤ tu piso rentable → quedas en el piso (sobre el mercado). Tu costo no da para competir.",
        "• En los 'No competitivos', bajar el margen baja tu piso, pero seguirás sobre el mercado: no te hace el más barato.",
        "• 'Bajan vs hoy' = productos que con ese margen quedarían MÁS BARATOS que su precio actual en Shopify (corrige sobreprecios).",
    ]
    for k, t in enumerate(notas):
        ws.cell(nr + 1 + k, 1, t)

    for col, w in {"A": 30, "B": 14, "C": 16, "D": 12, "E": 22, "F": 16, "G": 13, "H": 13}.items():
        ws.column_dimensions[col].width = w

    # --- Hoja Detalle ---
    wsd = wb.create_sheet("Detalle")
    cols = ["SKU", "Descripción", "Costo c/IVA", "Mediana mercado", "Fuentes",
            "Precio actual",
            "P. 22%", "Clasif 22%", "Mg% 22%",
            "P. 15%", "Clasif 15%", "Mg% 15%",
            "P. 10%", "Clasif 10%", "Mg% 10%"]
    for c, h in enumerate(cols, 1):
        cell = wsd.cell(1, c, h)
        cell.fill = azul
        cell.font = blanco_bold
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Ordenar: primero competitivos (donde la decisión importa), por margen potencial
    def orden(f):
        comp = f["Margen 15%_clasif"] == "Competitivo"
        return (0 if comp else 1, -f["costo_iva"])
    filas.sort(key=orden)

    for i, f in enumerate(filas, 2):
        wsd.cell(i, 1, f["sku"])
        wsd.cell(i, 2, f["desc"])
        wsd.cell(i, 3, int(f["costo_iva"]))
        wsd.cell(i, 4, int(f["mediana"]) if f["mediana"] else "—")
        wsd.cell(i, 5, f["fuentes"])
        wsd.cell(i, 6, int(f["precio_actual"]) if f["precio_actual"] else "—")
        col = 7
        for nombre_esc, _ in ESCENARIOS:
            wsd.cell(i, col, f[nombre_esc])
            clasif = f[nombre_esc + "_clasif"]
            cell_clasif = wsd.cell(i, col + 1, clasif)
            if clasif == "Competitivo":
                cell_clasif.fill = verde
            elif clasif == "No competitivo":
                cell_clasif.fill = amar
            else:
                cell_clasif.fill = rojo
            wsd.cell(i, col + 2, round(f[nombre_esc + "_margen"], 1))
            col += 3

    anchos_d = {"A": 14, "B": 40, "C": 12, "D": 14, "E": 8, "F": 12,
                "G": 10, "H": 13, "I": 9, "J": 10, "K": 13, "L": 9, "M": 10, "N": 13, "O": 9}
    for col, w in anchos_d.items():
        wsd.column_dimensions[col].width = w
    wsd.freeze_panes = "B2"

    # Guardar
    os.makedirs(DIR_OUT, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = os.path.join(DIR_OUT, f"simulacion_margenes_premium_{ts}.xlsx")
    wb.save(out)

    # Resumen por consola
    print("=" * 64)
    print("SIMULACIÓN DE MÁRGENES — ALTO COSTO")
    print("=" * 64)
    print(f"Productos alto costo: {len(filas)}\n")
    for nombre_esc, _ in ESCENARIOS:
        comp = sum(1 for f in filas if f[nombre_esc + "_clasif"] == "Competitivo")
        noc = sum(1 for f in filas if f[nombre_esc + "_clasif"] == "No competitivo")
        mono = sum(1 for f in filas if f[nombre_esc + "_clasif"] == "Monopolio")
        bajan = sum(1 for f in filas if f["precio_actual"] and f[nombre_esc] < f["precio_actual"])
        print(f"  {nombre_esc}: {comp} competitivos · {noc} no competitivos · {mono} monopolio · {bajan} bajan vs hoy")
    print(f"\n📊 Excel: {out}")
    print("\nRevisa la hoja 'Detalle': verde=competitivo (el margen cambia el precio),")
    print("amarillo=no competitivo (quedas en piso, tu costo no da para el mercado).")


if __name__ == "__main__":
    main()
